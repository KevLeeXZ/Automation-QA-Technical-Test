from openpyxl import load_workbook


def read_test_data(file_path):

    try:
        workbook = load_workbook(file_path)

        sheet = workbook["Test Data"]

        data = {}

        # Read header row (row 2)
        headers = {}

        for column in range(1, sheet.max_column + 1):
            header_value = sheet.cell(
                row=2,
                column=column
            ).value

            if header_value:
                headers[header_value] = column


        # Validate required columns exist
        if "Key" not in headers:
            raise Exception(
                "Excel file missing required column: Key"
            )

        if "Value" not in headers:
            raise Exception(
                "Excel file missing required column: Value"
            )


        key_column = headers["Key"]
        value_column = headers["Value"]


        # Read data starting from row 3
        for row in range(3, sheet.max_row + 1):

            key = sheet.cell(
                row=row,
                column=key_column
            ).value

            value = sheet.cell(
                row=row,
                column=value_column
            ).value


            if key:
                data[key] = value


        return data


    except Exception as e:

        raise Exception(
            f"Failed to read test data from '{file_path}'. "
            f"Error: {str(e)}"
        )